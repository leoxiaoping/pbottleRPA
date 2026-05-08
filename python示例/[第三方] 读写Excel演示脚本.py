"""
PBottle RPA demonstration demo. Please refer to the *process development documentation* for specific APIs.
Official website: https://rpa.pbottle.com/
Process development documentation: https://rpa.pbottle.com/docs/
"""

import sys
import os
import platform
import json
import time

# Add parent directory to path to import pbottleRPA
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import pbottleRPA as rpa

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    rpa.showMsg(
        "Please install the third-party module first", "Run: pip install openpyxl"
    )
    rpa.tts("Please install the third-party module first. Run: pip install openpyxl")
    rpa.exit("Please install the third-party module first. Run: pip install openpyxl")


def excel_append(filename, line=None):
    """
    Append a row of data to an Excel file.

    Args:
        filename (str): Absolute file path
        line (list): Row data
    """
    if line is None:
        line = []
    print("Appending row to Excel", filename, line)

    # Read the Excel file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Append a new row to the end of the existing sheet
    sheet.append(line)

    # Save
    workbook.save(filename)


print("=== Excel Read/Write Test ===")
print(rpa.getTime())
rpa.tts("Excel Read/Write Test")
rpa.wait(3)
rpa.tts("Generating an Excel file with current computer configuration information")
rpa.wait(5)

# Create an Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "pbottleRPA"

# Add header
sheet.append(["Item", "Value"])

# Make the first row bold
for cell in sheet[1]:
    cell.font = Font(bold=True)

# Set column widths
sheet.column_dimensions["A"].width = 30
sheet.column_dimensions["B"].width = 60

# Add data rows
sheet.append(
    [
        "Time",
        rpa.getTime(),
    ]
)
sheet.append(
    [
        "Display Resolution",
        json.dumps(rpa.getResolution()),
    ]
)
sheet.append(
    [
        "CPU",
        platform.processor() or "Unknown",
    ]
)

# Save the file
excel_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Excel_test_spreadsheet.xlsx",
)
workbook.save(excel_path)
rpa.tts("Excel test spreadsheet has been generated. Please check.")
rpa.openDir(os.path.dirname(excel_path))

# Append a row of data
excel_append(excel_path, ["Item name", "New appended value"])
rpa.wait(5)

# Read Excel
workbook2 = openpyxl.load_workbook(excel_path)
sheet2 = workbook2.active

# Retrieve all data
values = []
for row in sheet2.iter_rows(values_only=True):
    values.append(row)
print(values)
rpa.tts("Excel test spreadsheet has been read and logged.")


if __name__ == "__main__":
    pass
